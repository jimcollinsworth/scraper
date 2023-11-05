#import cProfile
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

#cProfile.run('main()', 'app.profile')
 # install snakeviz package
# generate report by calling "snakeviz app.profile"

def extract_urls_from_excel(file_path):
    """
    Extracts all URLs from an Excel file.
    
    Args:
    - file_path (str): path to the Excel file
    
    Returns:
    - list of URLs
    """
    urls = []
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    urls.append(cell.hyperlink.target)
    return urls

def extract_urls_from_bookmarks(file_path):
    """
    Extracts all URLs from Chrome's exported bookmarks HTML file.
    
    Args:
    - file_path (str): path to the exported bookmarks file
    
    Returns:
    - list of URLs
    """
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file.read(), 'html.parser')
    # Find all anchor tags and extract href values
    urls = [a['href'] for a in soup.find_all('a')]
    return urlsdef 
    
def extract_urls_from_txt(file_path):
    """
    Extracts all URLs from a txt or csv
    
    Args:
    - file_path (str): path to any csv, txt file 
    
    Returns:
    - list of URLs
    """
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    # Find all URLs in the content
    urls = re.findall(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', content)
    return urls

def save_urls_to_csv(urls, csv_path='urls.csv'):
    """
    Saves a list of URLs to a CSV file.
    
    Args:
    - urls (list): list of URLs
    - csv_path (str): path to save the CSV file
    """
    df = pd.DataFrame(urls, columns=['URLs'])
    df.to_csv(csv_path, index=False)

if __name__ == '__main__':
    file_type = input("Enter the file type (txt, html, or excel): ")
    if file_type == 'txt':
        file_path = input("Enter the path to the txt file: ")
        urls = extract_urls_from_txt(file_path)
    elif file_type == 'html':
        file_path = input("Enter the path to the bookmarks HTML file: ")
        urls = extract_urls_from_bookmarks(file_path)
    elif file_type == 'excel':
        file_path = input("Enter the path to the Excel file: ")
        urls = extract_urls_from_excel(file_path)
    else:
        print("Invalid file type. Please enter 'txt', 'html', or 'excel'.")
    
    output_csv_path = input("Enter the path to save the CSV file (or press enter for 'urls.csv'): ")
    if not output_csv_path:
        output_csv_path = 'urls.csv'
    
    save_urls_to_csv(urls, output_csv_path)
    print(f"URLs saved to {output_csv_path}")
